"""SweetControl - Backend (FastAPI)
Sistema de gestão para confeitaria.
"""
from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import io
import uuid
import logging
import bcrypt
import jwt
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Literal

from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, Query
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr, ConfigDict

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ============================================================================
# CONFIG
# ============================================================================
JWT_ALGORITHM = "HS256"
JWT_SECRET = os.environ["JWT_SECRET"]

# Cookie config: in production with cross-origin (Render frontend + Koyeb backend),
# browsers require SameSite=None + Secure=True to send cookies in cross-site XHRs.
COOKIE_SAMESITE = os.environ.get("COOKIE_SAMESITE", "lax").lower()  # "lax" | "none" | "strict"
COOKIE_SECURE = os.environ.get("COOKIE_SECURE", "false").lower() == "true"

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI(title="SweetControl API")
api = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger("sweetcontrol")


# ============================================================================
# AUTH HELPERS
# ============================================================================
def hash_password(p: str) -> str:
    return bcrypt.hashpw(p.encode(), bcrypt.gensalt()).decode()


def verify_password(p: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(p.encode(), h.encode())
    except Exception:
        return False


def create_access_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id, "email": email, "type": "access",
        "exp": datetime.now(timezone.utc) + timedelta(hours=12),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def create_refresh_token(user_id: str) -> str:
    payload = {"sub": user_id, "type": "refresh",
               "exp": datetime.now(timezone.utc) + timedelta(days=7)}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def set_auth_cookies(response: Response, user_id: str, email: str):
    access = create_access_token(user_id, email)
    refresh = create_refresh_token(user_id)
    response.set_cookie("access_token", access, httponly=True, secure=COOKIE_SECURE,
                        samesite=COOKIE_SAMESITE, max_age=43200, path="/")
    response.set_cookie("refresh_token", refresh, httponly=True, secure=COOKIE_SECURE,
                        samesite=COOKIE_SAMESITE, max_age=604800, path="/")


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Não autenticado")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Tipo de token inválido")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Usuário não encontrado")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")


# ============================================================================
# MODELS
# ============================================================================
def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def new_id() -> str:
    return str(uuid.uuid4())


class UserPublic(BaseModel):
    id: str
    email: str
    name: str
    role: str = "admin"
    created_at: str


class RegisterIn(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str


class LoginIn(BaseModel):
    email: EmailStr
    password: str


# ---- Customers ----
class CustomerIn(BaseModel):
    name: str
    phone: str = ""
    address: str = ""
    notes: str = ""


class Customer(CustomerIn):
    id: str
    created_at: str
    total_spent: float = 0.0
    order_count: int = 0


# ---- Ingredients ----
class IngredientIn(BaseModel):
    name: str
    unit: str = "g"  # g, kg, ml, L, un
    unit_price: float = 0.0  # price per unit
    stock_qty: float = 0.0
    min_stock: float = 0.0


class Ingredient(IngredientIn):
    id: str
    created_at: str


# ---- Doughs / Recipes (massa, recheio, cobertura) ----
class RecipeIngredient(BaseModel):
    ingredient_id: str
    ingredient_name: str = ""
    qty: float
    unit: str = "g"


class DoughIn(BaseModel):
    name: str
    category: Literal["massa", "recheio", "cobertura"] = "massa"
    ring_size: Optional[int] = None  # 10, 15, 20, 25, 30 (optional ref)
    yield_servings: float = 1.0
    ingredients: List[RecipeIngredient] = []
    notes: str = ""


class Dough(DoughIn):
    id: str
    total_cost: float = 0.0
    created_at: str


# ---- Price Table ----
class PriceRowIn(BaseModel):
    ring_size: int  # 10, 15, 20, 25, 30
    category: Literal["comum", "gourmet", "2_andares", "3_andares"] = "comum"
    cost: float = 0.0
    margin_percent: float = 100.0
    price: float = 0.0
    notes: str = ""


class PriceRow(PriceRowIn):
    id: str
    updated_at: str


# ---- Orders (encomendas) ----
class OrderIn(BaseModel):
    customer_name: str
    phone: str = ""
    address: str = ""
    delivery_date: str  # ISO date
    ring_size: int = 20
    dough: str = ""
    fillings: List[str] = []
    observations: str = ""
    total: float = 0.0
    status: Literal["pendente", "em_preparo", "finalizado", "entregue"] = "pendente"


class Order(OrderIn):
    id: str
    created_at: str


# ---- Sales (vendas) ----
class SaleIn(BaseModel):
    description: str
    ring_size: Optional[int] = None
    qty: int = 1
    unit_price: float
    cost: float = 0.0
    customer_name: str = ""
    sale_date: str  # ISO date


class Sale(SaleIn):
    id: str
    profit: float
    total: float
    created_at: str


# ---- Expenses (gastos) ----
class ExpenseIn(BaseModel):
    description: str
    category: str = "geral"
    amount: float
    expense_date: str


class Expense(ExpenseIn):
    id: str
    created_at: str


# ---- Settings ----
class Settings(BaseModel):
    business_name: str = "SweetControl"
    owner_name: str = ""
    address: str = ""
    phone: str = ""
    logo_url: str = ""
    default_margin: float = 100.0
    instagram: str = ""


# ---- Support Tickets ----
class SupportTicketIn(BaseModel):
    subject: str
    message: str
    category: Literal["bug", "duvida", "sugestao", "outro"] = "bug"
    priority: Literal["baixa", "media", "alta"] = "media"


class SupportReplyIn(BaseModel):
    message: str


class SupportTicket(BaseModel):
    id: str
    subject: str
    message: str
    category: str
    priority: str
    status: Literal["aberto", "em_andamento", "resolvido", "fechado"] = "aberto"
    author_name: str
    author_email: str
    replies: List[dict] = []
    created_at: str
    updated_at: str


# ============================================================================
# AUTH ROUTES
# ============================================================================
@api.post("/auth/register")
async def register(payload: RegisterIn, response: Response):
    email = payload.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email já cadastrado")
    user = {
        "id": new_id(),
        "email": email,
        "name": payload.name,
        "password_hash": hash_password(payload.password),
        "role": "admin",
        "created_at": now_iso(),
    }
    await db.users.insert_one(user)
    set_auth_cookies(response, user["id"], user["email"])
    return {"id": user["id"], "email": user["email"], "name": user["name"], "role": user["role"], "created_at": user["created_at"]}


@api.post("/auth/login")
async def login(payload: LoginIn, response: Response):
    email = payload.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Email ou senha incorretos")
    set_auth_cookies(response, user["id"], user["email"])
    return {"id": user["id"], "email": user["email"], "name": user["name"], "role": user.get("role", "admin"), "created_at": user["created_at"]}


@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"ok": True}


@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user


# ============================================================================
# CUSTOMERS
# ============================================================================
@api.get("/customers", response_model=List[Customer])
async def list_customers(user=Depends(get_current_user)):
    items = await db.customers.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)
    # enrich with totals
    for c in items:
        sales_cursor = db.orders.find({"customer_name": c["name"]}, {"_id": 0, "total": 1})
        total = 0.0
        cnt = 0
        async for s in sales_cursor:
            total += float(s.get("total", 0))
            cnt += 1
        c["total_spent"] = round(total, 2)
        c["order_count"] = cnt
    return items


@api.post("/customers", response_model=Customer)
async def create_customer(payload: CustomerIn, user=Depends(get_current_user)):
    doc = {**payload.model_dump(), "id": new_id(), "created_at": now_iso(),
           "total_spent": 0.0, "order_count": 0}
    await db.customers.insert_one(doc.copy())
    doc.pop("_id", None)
    return doc


@api.put("/customers/{cid}", response_model=Customer)
async def update_customer(cid: str, payload: CustomerIn, user=Depends(get_current_user)):
    await db.customers.update_one({"id": cid}, {"$set": payload.model_dump()})
    c = await db.customers.find_one({"id": cid}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    c.setdefault("total_spent", 0.0)
    c.setdefault("order_count", 0)
    return c


@api.delete("/customers/{cid}")
async def delete_customer(cid: str, user=Depends(get_current_user)):
    await db.customers.delete_one({"id": cid})
    return {"ok": True}


# ============================================================================
# INGREDIENTS / ESTOQUE
# ============================================================================
@api.get("/ingredients", response_model=List[Ingredient])
async def list_ingredients(user=Depends(get_current_user)):
    return await db.ingredients.find({}, {"_id": 0}).sort("name", 1).to_list(5000)


@api.post("/ingredients", response_model=Ingredient)
async def create_ingredient(payload: IngredientIn, user=Depends(get_current_user)):
    doc = {**payload.model_dump(), "id": new_id(), "created_at": now_iso()}
    await db.ingredients.insert_one(doc.copy())
    doc.pop("_id", None)
    return doc


@api.put("/ingredients/{iid}", response_model=Ingredient)
async def update_ingredient(iid: str, payload: IngredientIn, user=Depends(get_current_user)):
    await db.ingredients.update_one({"id": iid}, {"$set": payload.model_dump()})
    item = await db.ingredients.find_one({"id": iid}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Ingrediente não encontrado")
    return item


@api.delete("/ingredients/{iid}")
async def delete_ingredient(iid: str, user=Depends(get_current_user)):
    await db.ingredients.delete_one({"id": iid})
    return {"ok": True}


class StockMovementIn(BaseModel):
    ingredient_id: str
    delta: float  # positive = entrada, negative = saída
    reason: str = ""


@api.post("/ingredients/movement")
async def stock_movement(payload: StockMovementIn, user=Depends(get_current_user)):
    ing = await db.ingredients.find_one({"id": payload.ingredient_id}, {"_id": 0})
    if not ing:
        raise HTTPException(status_code=404, detail="Ingrediente não encontrado")
    new_qty = float(ing.get("stock_qty", 0)) + payload.delta
    await db.ingredients.update_one({"id": payload.ingredient_id}, {"$set": {"stock_qty": new_qty}})
    await db.stock_movements.insert_one({
        "id": new_id(), "ingredient_id": payload.ingredient_id,
        "ingredient_name": ing["name"], "delta": payload.delta,
        "reason": payload.reason, "created_at": now_iso(),
    })
    return {"ok": True, "stock_qty": new_qty}


@api.get("/stock-movements")
async def list_movements(user=Depends(get_current_user)):
    return await db.stock_movements.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)


# ============================================================================
# DOUGHS (MASSAS / RECHEIOS)
# ============================================================================
async def compute_dough_cost(ingredients: List[dict]) -> float:
    total = 0.0
    for ri in ingredients:
        ing = await db.ingredients.find_one({"id": ri["ingredient_id"]}, {"_id": 0})
        if ing:
            total += float(ing.get("unit_price", 0)) * float(ri.get("qty", 0))
    return round(total, 2)


@api.get("/doughs", response_model=List[Dough])
async def list_doughs(user=Depends(get_current_user)):
    return await db.doughs.find({}, {"_id": 0}).sort("name", 1).to_list(2000)


@api.post("/doughs", response_model=Dough)
async def create_dough(payload: DoughIn, user=Depends(get_current_user)):
    data = payload.model_dump()
    data["total_cost"] = await compute_dough_cost(data["ingredients"])
    doc = {**data, "id": new_id(), "created_at": now_iso()}
    await db.doughs.insert_one(doc.copy())
    doc.pop("_id", None)
    return doc


@api.put("/doughs/{did}", response_model=Dough)
async def update_dough(did: str, payload: DoughIn, user=Depends(get_current_user)):
    data = payload.model_dump()
    data["total_cost"] = await compute_dough_cost(data["ingredients"])
    await db.doughs.update_one({"id": did}, {"$set": data})
    doc = await db.doughs.find_one({"id": did}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Massa não encontrada")
    return doc


@api.delete("/doughs/{did}")
async def delete_dough(did: str, user=Depends(get_current_user)):
    await db.doughs.delete_one({"id": did})
    return {"ok": True}


# ============================================================================
# PRICES
# ============================================================================
@api.get("/prices", response_model=List[PriceRow])
async def list_prices(user=Depends(get_current_user)):
    return await db.prices.find({}, {"_id": 0}).sort([("ring_size", 1), ("category", 1)]).to_list(500)


@api.post("/prices", response_model=PriceRow)
async def create_price(payload: PriceRowIn, user=Depends(get_current_user)):
    data = payload.model_dump()
    # auto compute price = cost * (1 + margin/100) if price=0
    if not data.get("price"):
        data["price"] = round(data["cost"] * (1 + data["margin_percent"] / 100.0), 2)
    doc = {**data, "id": new_id(), "updated_at": now_iso()}
    await db.prices.insert_one(doc.copy())
    doc.pop("_id", None)
    return doc


@api.put("/prices/{pid}", response_model=PriceRow)
async def update_price(pid: str, payload: PriceRowIn, user=Depends(get_current_user)):
    data = payload.model_dump()
    if not data.get("price"):
        data["price"] = round(data["cost"] * (1 + data["margin_percent"] / 100.0), 2)
    data["updated_at"] = now_iso()
    await db.prices.update_one({"id": pid}, {"$set": data})
    doc = await db.prices.find_one({"id": pid}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Preço não encontrado")
    return doc


@api.delete("/prices/{pid}")
async def delete_price(pid: str, user=Depends(get_current_user)):
    await db.prices.delete_one({"id": pid})
    return {"ok": True}


# ============================================================================
# ORDERS
# ============================================================================
@api.get("/orders", response_model=List[Order])
async def list_orders(status: Optional[str] = None, q: Optional[str] = None,
                      user=Depends(get_current_user)):
    query = {}
    if status:
        query["status"] = status
    if q:
        query["customer_name"] = {"$regex": q, "$options": "i"}
    return await db.orders.find(query, {"_id": 0}).sort("delivery_date", 1).to_list(2000)


@api.post("/orders", response_model=Order)
async def create_order(payload: OrderIn, user=Depends(get_current_user)):
    doc = {**payload.model_dump(), "id": new_id(), "created_at": now_iso()}
    await db.orders.insert_one(doc.copy())
    doc.pop("_id", None)
    # auto-create customer if doesn't exist
    if payload.customer_name:
        exists = await db.customers.find_one({"name": payload.customer_name})
        if not exists:
            await db.customers.insert_one({
                "id": new_id(), "name": payload.customer_name,
                "phone": payload.phone, "address": payload.address, "notes": "",
                "created_at": now_iso(), "total_spent": 0.0, "order_count": 0,
            })
    return doc


@api.put("/orders/{oid}", response_model=Order)
async def update_order(oid: str, payload: OrderIn, user=Depends(get_current_user)):
    await db.orders.update_one({"id": oid}, {"$set": payload.model_dump()})
    doc = await db.orders.find_one({"id": oid}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Encomenda não encontrada")
    return doc


@api.delete("/orders/{oid}")
async def delete_order(oid: str, user=Depends(get_current_user)):
    await db.orders.delete_one({"id": oid})
    return {"ok": True}


# ============================================================================
# SALES
# ============================================================================
@api.get("/sales", response_model=List[Sale])
async def list_sales(user=Depends(get_current_user)):
    return await db.sales.find({}, {"_id": 0}).sort("sale_date", -1).to_list(5000)


@api.post("/sales", response_model=Sale)
async def create_sale(payload: SaleIn, user=Depends(get_current_user)):
    data = payload.model_dump()
    total = round(data["unit_price"] * data["qty"], 2)
    profit = round(total - (data["cost"] * data["qty"]), 2)
    doc = {**data, "id": new_id(), "total": total, "profit": profit, "created_at": now_iso()}
    await db.sales.insert_one(doc.copy())
    doc.pop("_id", None)
    return doc


@api.delete("/sales/{sid}")
async def delete_sale(sid: str, user=Depends(get_current_user)):
    await db.sales.delete_one({"id": sid})
    return {"ok": True}


# ============================================================================
# EXPENSES
# ============================================================================
@api.get("/expenses", response_model=List[Expense])
async def list_expenses(user=Depends(get_current_user)):
    return await db.expenses.find({}, {"_id": 0}).sort("expense_date", -1).to_list(5000)


@api.post("/expenses", response_model=Expense)
async def create_expense(payload: ExpenseIn, user=Depends(get_current_user)):
    doc = {**payload.model_dump(), "id": new_id(), "created_at": now_iso()}
    await db.expenses.insert_one(doc.copy())
    doc.pop("_id", None)
    return doc


@api.delete("/expenses/{eid}")
async def delete_expense(eid: str, user=Depends(get_current_user)):
    await db.expenses.delete_one({"id": eid})
    return {"ok": True}


# ============================================================================
# DASHBOARD
# ============================================================================
@api.get("/dashboard")
async def dashboard(user=Depends(get_current_user)):
    today = datetime.now(timezone.utc).date().isoformat()
    month_prefix = today[:7]

    sales = await db.sales.find({}, {"_id": 0}).to_list(10000)
    expenses = await db.expenses.find({}, {"_id": 0}).to_list(10000)
    orders = await db.orders.find({}, {"_id": 0}).to_list(10000)

    total_day = sum(s.get("total", 0) for s in sales if s.get("sale_date", "").startswith(today))
    total_month = sum(s.get("total", 0) for s in sales if s.get("sale_date", "").startswith(month_prefix))
    profit_month = sum(s.get("profit", 0) for s in sales if s.get("sale_date", "").startswith(month_prefix))
    expenses_month = sum(e.get("amount", 0) for e in expenses if e.get("expense_date", "").startswith(month_prefix))
    pending_orders = len([o for o in orders if o.get("status") in ("pendente", "em_preparo")])

    # Top products
    prod_map = {}
    for s in sales:
        k = s.get("description", "Outros")
        prod_map[k] = prod_map.get(k, 0) + s.get("qty", 1)
    top_products = sorted(
        [{"name": k, "qty": v} for k, v in prod_map.items()],
        key=lambda x: x["qty"], reverse=True
    )[:5]

    # Last 7 days chart
    sales_chart = []
    expenses_chart = []
    profit_chart = []
    for i in range(6, -1, -1):
        d = (datetime.now(timezone.utc) - timedelta(days=i)).date().isoformat()
        day_sales = sum(s.get("total", 0) for s in sales if s.get("sale_date", "").startswith(d))
        day_profit = sum(s.get("profit", 0) for s in sales if s.get("sale_date", "").startswith(d))
        day_exp = sum(e.get("amount", 0) for e in expenses if e.get("expense_date", "").startswith(d))
        label = datetime.fromisoformat(d).strftime("%d/%m")
        sales_chart.append({"day": label, "value": round(day_sales, 2)})
        expenses_chart.append({"day": label, "value": round(day_exp, 2)})
        profit_chart.append({"day": label, "value": round(day_profit, 2)})

    # Low stock alerts
    low_stock = []
    async for ing in db.ingredients.find({}, {"_id": 0}):
        if float(ing.get("stock_qty", 0)) <= float(ing.get("min_stock", 0)) and float(ing.get("min_stock", 0)) > 0:
            low_stock.append({"name": ing["name"], "stock_qty": ing["stock_qty"],
                              "min_stock": ing["min_stock"], "unit": ing.get("unit", "")})

    return {
        "total_day": round(total_day, 2),
        "total_month": round(total_month, 2),
        "profit_month": round(profit_month, 2),
        "expenses_month": round(expenses_month, 2),
        "pending_orders": pending_orders,
        "total_orders": len(orders),
        "top_products": top_products,
        "sales_chart": sales_chart,
        "expenses_chart": expenses_chart,
        "profit_chart": profit_chart,
        "low_stock": low_stock,
    }


# ============================================================================
# SETTINGS
# ============================================================================
@api.get("/settings", response_model=Settings)
async def get_settings(user=Depends(get_current_user)):
    s = await db.settings.find_one({"id": "main"}, {"_id": 0})
    if not s:
        default = Settings().model_dump()
        await db.settings.insert_one({**default, "id": "main"})
        return default
    s.pop("id", None)
    return s


@api.put("/settings", response_model=Settings)
async def update_settings(payload: Settings, user=Depends(get_current_user)):
    await db.settings.update_one({"id": "main"}, {"$set": payload.model_dump()}, upsert=True)
    return payload


# ============================================================================
# SUPPORT TICKETS
# ============================================================================
@api.get("/support", response_model=List[SupportTicket])
async def list_tickets(user=Depends(get_current_user)):
    return await db.support_tickets.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)


@api.post("/support", response_model=SupportTicket)
async def create_ticket(payload: SupportTicketIn, user=Depends(get_current_user)):
    now = now_iso()
    doc = {
        **payload.model_dump(),
        "id": new_id(),
        "status": "aberto",
        "author_name": user.get("name", ""),
        "author_email": user.get("email", ""),
        "replies": [],
        "created_at": now,
        "updated_at": now,
    }
    await db.support_tickets.insert_one(doc.copy())
    doc.pop("_id", None)
    return doc


@api.post("/support/{tid}/reply", response_model=SupportTicket)
async def reply_ticket(tid: str, payload: SupportReplyIn, user=Depends(get_current_user)):
    ticket = await db.support_tickets.find_one({"id": tid}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket não encontrado")
    reply = {
        "id": new_id(),
        "author_name": user.get("name", ""),
        "author_email": user.get("email", ""),
        "message": payload.message,
        "created_at": now_iso(),
    }
    replies = ticket.get("replies", []) + [reply]
    await db.support_tickets.update_one(
        {"id": tid},
        {"$set": {"replies": replies, "updated_at": now_iso(), "status": "em_andamento"}},
    )
    ticket = await db.support_tickets.find_one({"id": tid}, {"_id": 0})
    return ticket


class SupportStatusIn(BaseModel):
    status: Literal["aberto", "em_andamento", "resolvido", "fechado"]


@api.put("/support/{tid}/status", response_model=SupportTicket)
async def update_ticket_status(tid: str, payload: SupportStatusIn, user=Depends(get_current_user)):
    await db.support_tickets.update_one(
        {"id": tid}, {"$set": {"status": payload.status, "updated_at": now_iso()}}
    )
    ticket = await db.support_tickets.find_one({"id": tid}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket não encontrado")
    return ticket


@api.delete("/support/{tid}")
async def delete_ticket(tid: str, user=Depends(get_current_user)):
    await db.support_tickets.delete_one({"id": tid})
    return {"ok": True}


# ============================================================================
# REPORTS - PDF / EXCEL
# ============================================================================
def _build_pdf(title: str, headers: List[str], rows: List[List[str]]) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=20,
                                  textColor=colors.HexColor("#4A2D35"), spaceAfter=12)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], fontSize=10,
                                textColor=colors.HexColor("#8A8581"), spaceAfter=20)
    elements = [
        Paragraph(f"<b>SweetControl</b> — {title}", title_style),
        Paragraph(f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}", sub_style),
    ]
    data = [headers] + rows
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E4B5C6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#4A2D35")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FDFBF7")),
        ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#2D2A26")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#FDFBF7"), colors.HexColor("#FFFFFF")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#F0EAE1")),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 8),
        ("TOPPADDING", (0, 1), (-1, -1), 8),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return buf.read()


def _build_xlsx(title: str, headers: List[str], rows: List[List]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    header_fill = PatternFill("solid", fgColor="E4B5C6")
    bold = Font(bold=True, color="4A2D35")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@api.get("/reports/sales/{fmt}")
async def report_sales(fmt: str, user=Depends(get_current_user)):
    sales = await db.sales.find({}, {"_id": 0}).sort("sale_date", -1).to_list(5000)
    headers = ["Data", "Descrição", "Aro", "Qtd", "Preço Un.", "Custo", "Total", "Lucro"]
    rows = [[s.get("sale_date", ""), s.get("description", ""), s.get("ring_size") or "-",
             s.get("qty", 0), f"R$ {s.get('unit_price', 0):.2f}", f"R$ {s.get('cost', 0):.2f}",
             f"R$ {s.get('total', 0):.2f}", f"R$ {s.get('profit', 0):.2f}"] for s in sales]
    if fmt == "pdf":
        data = _build_pdf("Relatório de Vendas", headers, rows)
        return StreamingResponse(io.BytesIO(data), media_type="application/pdf",
                                 headers={"Content-Disposition": "attachment; filename=vendas.pdf"})
    elif fmt == "xlsx":
        data = _build_xlsx("Vendas", headers, rows)
        return StreamingResponse(io.BytesIO(data),
                                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=vendas.xlsx"})
    raise HTTPException(status_code=400, detail="Formato inválido")


@api.get("/reports/expenses/{fmt}")
async def report_expenses(fmt: str, user=Depends(get_current_user)):
    items = await db.expenses.find({}, {"_id": 0}).sort("expense_date", -1).to_list(5000)
    headers = ["Data", "Descrição", "Categoria", "Valor"]
    rows = [[i.get("expense_date", ""), i.get("description", ""), i.get("category", ""),
             f"R$ {i.get('amount', 0):.2f}"] for i in items]
    if fmt == "pdf":
        data = _build_pdf("Relatório de Gastos", headers, rows)
        return StreamingResponse(io.BytesIO(data), media_type="application/pdf",
                                 headers={"Content-Disposition": "attachment; filename=gastos.pdf"})
    elif fmt == "xlsx":
        data = _build_xlsx("Gastos", headers, rows)
        return StreamingResponse(io.BytesIO(data),
                                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=gastos.xlsx"})
    raise HTTPException(status_code=400, detail="Formato inválido")


@api.get("/reports/orders/{fmt}")
async def report_orders(fmt: str, user=Depends(get_current_user)):
    items = await db.orders.find({}, {"_id": 0}).sort("delivery_date", -1).to_list(5000)
    headers = ["Entrega", "Cliente", "Telefone", "Aro", "Massa", "Status", "Total"]
    rows = [[o.get("delivery_date", ""), o.get("customer_name", ""), o.get("phone", ""),
             o.get("ring_size", ""), o.get("dough", ""), o.get("status", ""),
             f"R$ {o.get('total', 0):.2f}"] for o in items]
    if fmt == "pdf":
        data = _build_pdf("Relatório de Encomendas", headers, rows)
        return StreamingResponse(io.BytesIO(data), media_type="application/pdf",
                                 headers={"Content-Disposition": "attachment; filename=encomendas.pdf"})
    elif fmt == "xlsx":
        data = _build_xlsx("Encomendas", headers, rows)
        return StreamingResponse(io.BytesIO(data),
                                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=encomendas.xlsx"})
    raise HTTPException(status_code=400, detail="Formato inválido")


# ============================================================================
# HEALTH
# ============================================================================
@api.get("/")
async def root():
    return {"app": "SweetControl", "status": "ok"}


# ============================================================================
# STARTUP - SEED
# ============================================================================
@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.customers.create_index("name")
    await db.ingredients.create_index("name")
    # seed admin
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@sweetcontrol.com").lower()
    admin_pass = os.environ.get("ADMIN_PASSWORD", "sweet123")
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({
            "id": new_id(), "email": admin_email, "name": "Admin",
            "password_hash": hash_password(admin_pass), "role": "admin",
            "created_at": now_iso(),
        })
        logger.info(f"Admin seeded: {admin_email}")
    elif not verify_password(admin_pass, existing["password_hash"]):
        await db.users.update_one({"email": admin_email},
                                  {"$set": {"password_hash": hash_password(admin_pass)}})
        logger.info(f"Admin password updated: {admin_email}")

    # seed sample ingredients if empty
    if await db.ingredients.count_documents({}) == 0:
        seed_ings = [
            {"name": "Ovos", "unit": "un", "unit_price": 0.80, "stock_qty": 60, "min_stock": 12},
            {"name": "Leite Integral", "unit": "ml", "unit_price": 0.006, "stock_qty": 5000, "min_stock": 1000},
            {"name": "Farinha de Trigo", "unit": "g", "unit_price": 0.005, "stock_qty": 5000, "min_stock": 1000},
            {"name": "Açúcar Refinado", "unit": "g", "unit_price": 0.004, "stock_qty": 5000, "min_stock": 1000},
            {"name": "Óleo", "unit": "ml", "unit_price": 0.008, "stock_qty": 2000, "min_stock": 500},
            {"name": "Fermento em Pó", "unit": "g", "unit_price": 0.06, "stock_qty": 500, "min_stock": 100},
            {"name": "Chocolate em Pó", "unit": "g", "unit_price": 0.03, "stock_qty": 1000, "min_stock": 200},
            {"name": "Leite Condensado", "unit": "g", "unit_price": 0.015, "stock_qty": 3950, "min_stock": 790},
            {"name": "Manteiga", "unit": "g", "unit_price": 0.04, "stock_qty": 1000, "min_stock": 200},
            {"name": "Achocolatado", "unit": "g", "unit_price": 0.025, "stock_qty": 1000, "min_stock": 200},
        ]
        now = now_iso()
        await db.ingredients.insert_many([
            {**i, "id": new_id(), "created_at": now} for i in seed_ings
        ])
        logger.info("Sample ingredients seeded")

    # seed default settings
    if not await db.settings.find_one({"id": "main"}):
        s = Settings().model_dump()
        await db.settings.insert_one({**s, "id": "main"})


@app.on_event("shutdown")
async def shutdown():
    client.close()


# Mount router
app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)
